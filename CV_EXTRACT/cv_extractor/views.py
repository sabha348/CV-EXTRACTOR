import os
import uuid
import re
import pandas as pd
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, FileResponse
from django.conf import settings
from PyPDF2 import PdfReader
import docx
import pythoncom
from win32com import client
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def home(request):
    return redirect('upload_cv')


def sanitize_text(text):
    # Replace or remove illegal characters as necessary
    # Remove characters that are not allowed by openpyxl
    illegal_chars = ['\r', '\n', '\t', chr(11), chr(12), chr(14), chr(15), chr(19), chr(21), chr(127)]
    # Additional illegal characters based on openpyxl documentation
    illegal_chars += [chr(i) for i in range(32)]
    illegal_chars += ['\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08', '\x0B', '\x0C', '\x0E', '\x0F', '\x10', '\x11', '\x12', '\x13', '\x14', '\x15', '\x16', '\x17', '\x18', '\x19', '\x1A', '\x1B', '\x1C', '\x1D', '\x1E', '\x1F', '\x7F']
    for char in illegal_chars:
        text = text.replace(char, '')
        # New line to add space before and after email address
    text = re.sub(r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,})', r' \1 ', text)
    return text

def upload_cv(request):
    if request.method == 'POST':
        uploaded_files = request.FILES.getlist('document')
        if uploaded_files:
            # Initialize COM libraries
            pythoncom.CoInitialize()
            word = client.Dispatch("Word.Application")

            # Create a temp directory inside MEDIA_ROOT if it doesn't exist
            temp_dir = os.path.join(settings.MEDIA_ROOT, 'cv_extractor', 'temp')
            os.makedirs(temp_dir, exist_ok=True)

            # Initialize an empty DataFrame
            all_data = pd.DataFrame(columns=['Email', 'Contact Number', 'Overall Text'])

             # Initialize the list to keep track of temporary file paths
            temp_files = []

            for uploaded_file in uploaded_files:
                # Process each file as before
                # Save the uploaded file to the temp directory
                temp_file_name = str(uuid.uuid4())  # Generate a unique file name
                temp_file_extension = os.path.splitext(uploaded_file.name)[1]
                temp_file_path = os.path.join(temp_dir, temp_file_name + temp_file_extension)
                temp_files.append(temp_file_path)  # Add the temp file path to the list
                with open(temp_file_path, 'wb+') as temp_file:
                    for chunk in uploaded_file.chunks():
                        temp_file.write(chunk)

                # Process the file based on its extension
                text = ""
                if uploaded_file.name.endswith('.pdf'):
                    reader = PdfReader(temp_file_path)
                    text = "\n".join([page.extract_text() for page in reader.pages])
                elif uploaded_file.name.endswith('.docx'):
                    doc = docx.Document(temp_file_path)
                    text = "\n".join([para.text for para in doc.paragraphs])
                elif uploaded_file.name.endswith('.doc'):
                    doc = word.Documents.Open(temp_file_path)
                    text = doc.Range().Text
                    doc.Close(False)  # Close the document without saving changes
                else:
                    # Uninitialize COM libraries
                    pythoncom.CoUninitialize()
                    # Clean up the temporary file
                    os.remove(temp_file_path)
                    return HttpResponse("Unsupported file type.", status=400)

                # Sanitize text to remove illegal characters
                text = sanitize_text(text)

                # Find email and contact number using regex
                # Updated regex patterns
                # Updated regex pattern for email extraction
                email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                # Updated regex pattern for contact number extraction
                contact_pattern = r'\b\d{5}\s?\d{5}\b'

                emails = re.findall(email_pattern, text)
                contacts = re.findall(contact_pattern, text)

                # Filter out any email that starts with a number sequence that could be a phone number
                emails = [email for email in emails if not re.match(r'^\d{10}', email)]

                # Additional processing to isolate email address
                # This is a simple heuristic that assumes the email is followed by certain keywords
                keywords = ['CAREEROBJECTIVE', 'RESUME', 'CV', 'OBJECTIVE', 'PROFILE']
                for keyword in keywords:
                    emails = [email.split(keyword)[0] for email in emails]

                # Isolate email address by splitting text at common keywords or dates
                for keyword in ['Date of birth', 'Nationality', 'SKILLS', 'PROFILE']:
                    emails = [email.split(keyword)[0].strip() for email in emails]
                # Join multiple emails and contact numbers into a single string
                # Join multiple emails and contact numbers into a single string
                email_str = ', '.join(set(emails))  # Use set to remove duplicates
                contact_str = ', '.join(set(contacts))

                # After extracting text and emails/contacts
                text = text.replace('\n', ' ')  # Replace newline characters with a space

            # Append the data to the all_data DataFrame instead of creating a new one
                all_data = pd.concat([all_data, pd.DataFrame([{
                'Email': email_str,
                'Contact Number': contact_str,
                'Overall Text': text
                }])], ignore_index=True)

            # After processing all files
            # Generate a unique file name for the Excel file
            unique_filename = str(uuid.uuid4()) + '.xlsx'
            excel_file_path = os.path.join(settings.MEDIA_ROOT, 'cv_extractor', 'xlsx', unique_filename)
            # After df.to_excel(excel_file_path, engine='openpyxl')
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

             # Convert the all_data DataFrame to Excel file
            all_data.to_excel(excel_file_path, index=False, engine='openpyxl')

            # Load the workbook and select the active worksheet
            wb = load_workbook(excel_file_path)
            ws = wb.active

            # Iterate over all cells in the DataFrame to adjust the column width and row height
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

            for row_cells in ws.iter_rows():
                height = max(cell.value.count('\n') + 1 for cell in row_cells if cell.value)
                ws.row_dimensions[row_cells[0].row].height = height * 15  # Adjust multiplier as needed for row height

            # Set the width of the column for 'Overall Text' to a fixed width
            ws.column_dimensions['C'].width = 50  # Adjust the column width as needed

            # Set text wrapping and alignment for the 'Overall Text' cell
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # Assuming 'Overall Text' is in the third column
                for cell in row:
                    cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')

            # Save the updated workbook
            wb.save(excel_file_path)

            # Clean up the temporary files after processing
            for temp_file_path in temp_files:
                os.remove(temp_file_path)

            # Uninitialize COM libraries
            word.Quit()
            pythoncom.CoUninitialize()

            # Return the path to the newly created Excel file
            return JsonResponse({'file_path': unique_filename})
        else:
            return HttpResponse("No files were uploaded.", status=400)
    else:
        return render(request, 'cv_extractor/upload_form.html')

def download_cv(request, file_path):
    # Ensure the file_path includes the correct subdirectories and file extension
    file_full_path = os.path.join(settings.MEDIA_ROOT, 'cv_extractor', 'xlsx', file_path)
    try:
        # Attempt to open the file and return it as a response
        return FileResponse(open(file_full_path, 'rb'), as_attachment=True, filename=file_path)
    except FileNotFoundError:
        # If the file is not found, return an appropriate response
        return HttpResponse("The requested file does not exist.", status=404)